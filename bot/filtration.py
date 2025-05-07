from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium import webdriver
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from threading import Event
import openpyxl
import os
import re
import time
import sys

class BookingFiltration:
    
    def __init__(self, driver: WebDriver):
        self.driver = driver
        
        
        self.driver.set_script_timeout(30000)
        self.excel_file = self._get_excel_filename()
        if not os.path.exists(self.excel_file):
            self._initialize_excel()
        
        # Add a counter for processed products
        self.processed_products = 0
        # Add a lock for thread-safe Excel operations
        from threading import Lock
        self.excel_lock = Lock()
        self.stop_event = Event()


    
    def get_progress_percentage(self):
        """Calculate the progress percentage based on processed vs total products"""
        if hasattr(self, 'total_products') and self.total_products > 0:
            return (self.processed_products / self.total_products) * 100
        return 0

    def set_progress_callback(self, callback):
        """Set a callback function to be called when progress updates"""
        self.progress_callback = callback
        
    def _get_excel_filename(self):
        try:
            main_title = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "main-title"))
            ).text.strip()

            sanitized_title = re.sub(r"[^\w\s-]", "", main_title).replace(" ", "_")
            return f"products_{sanitized_title}.xlsx"
        except TimeoutException:
            print("Main title not found. Using default file name.")
            return "products.xlsx"

    def _initialize_excel(self):
        # Get the folder where the executable is located
        if getattr(sys, 'frozen', False):
            # When running as a packaged executable (PyInstaller)
            executable_dir = os.path.dirname(sys.executable)
        else:
            # When running as a normal script
            executable_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Define the path for the "produtos" folder
        produtos_folder = os.path.join(executable_dir, "produtos")
        
        # Create the "produtos" folder if it doesn't exist
        if not os.path.exists(produtos_folder):
            os.makedirs(produtos_folder)

        # Ensure that self.excel_file is a valid file name (correct for special characters)
        excel_filename = self.excel_file.replace('/', '_').replace(':', '_')  # Sanitize file name
        excel_path = os.path.join(produtos_folder, excel_filename)
        
        # Create a new workbook and set up the sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["Título", "Code", "Código de barras","Marca", "preço atual","Preço Original ","Preço Assinante","Estoque", "Variação"])
        
        # Attempt to save the workbook
        try:
            workbook.save(excel_path)
            print(f"Excel file saved successfully at {excel_path}")
        except Exception as e:
            print(f"Failed to save to Excel: {e}")

    def _append_to_excel(self, data):
        """Thread-safe method to append data to Excel"""
        with self.excel_lock:  # Ensure thread safety using the lock
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    # Ensure the file path is correct, especially with PyInstaller
                    if getattr(sys, 'frozen', False):
                        executable_dir = os.path.dirname(sys.executable)
                    else:
                        executable_dir = os.path.dirname(os.path.abspath(__file__))
                    
                    # Define the path for the "produtos" folder
                    produtos_folder = os.path.join(executable_dir, "produtos")
                    
                    # Create the folder if it doesn't exist
                    if not os.path.exists(produtos_folder):
                        os.makedirs(produtos_folder)
                    
                    # Sanitize the filename to remove special characters
                    excel_filename = self.excel_file.replace('/', '_').replace(':', '_').replace('ç', 'c').replace('á', 'a').replace('ã', 'a')
                    excel_path = os.path.join(produtos_folder, excel_filename)

                    # If the file exists, open and append data
                    if os.path.exists(excel_path):
                        workbook = openpyxl.load_workbook(excel_path)
                    else:
                        # If the file doesn't exist, create a new one
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        sheet.title = "Products"
                        sheet.append(["Título", "Code", "Código de barras","Marca", "Preço atual","Preço Original","Preço assinante","Estoque", "Variação"])
                    
                    sheet = workbook.active
                    sheet.append(data)
                    workbook.save(excel_path)

                    self.processed_products += 1
                    if hasattr(self, 'progress_callback'):
                        progress = self.get_progress_percentage()
                        self.progress_callback(progress)
                    print(f"Successfully saved product {self.processed_products}")
                    break

                except Exception as e:
                    if attempt == max_retries - 1:
                        print(f"Failed to save to Excel after {max_retries} attempts: {e}")
                    else:
                        print(f"Error appending data, retrying... ({attempt + 1}/{max_retries})")
                        time.sleep(1)

    def filter_products(self):
        """Retrieve all product links and process them in parallel."""
        self._load_all_products()
        if self.stop_event.is_set():
            return
            
        product_links = self._extract_links_from_current_page()
        print(f"Found {len(product_links)} products to process")
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = []
            for link in product_links:
                if self.stop_event.is_set():
                    break
                futures.append(executor.submit(self._process_product_in_new_instance, link))

            for future in as_completed(futures):
                if self.stop_event.is_set():
                    executor.shutdown(wait=False)
                    break
                try:
                    future.result()
                except Exception as e:
                    print(f"Error processing product: {e}")
                    
        print(f"Finished processing all products. Total processed: {self.processed_products}")
    """
    def _load_all_products(self):
        #Modified to ensure all products are loaded
        print("Starting to load all products...")
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        
        while True:
            # Scroll down to bottom
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            
            # Wait for new elements to load
            time.sleep(2)
            
            # Calculate new scroll height and compare with last scroll height
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                # Try clicking "Load More" button if it exists
                try:
                    load_more = self.driver.find_element(By.ID, "next")
                    if load_more.is_displayed():
                        load_more.click()
                        time.sleep(2)
                    else:
                        break
                except:
                    break
            last_height = new_height
        
        print("Finished loading all products")
    """
    
    
    def _load_all_products(self):
        print("Starting to load all products...")
        
        # Aguardar explicitamente pelo container de produtos
        try:
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.card-product.card-product-showcase, .product-card"))
            )
        except TimeoutException:
            print("Timeout esperando produtos carregarem")
            return
        
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_attempts = 10
        
        while scroll_attempts < max_attempts and not self.stop_event.is_set():
            # Scroll em partes menores
            self.driver.execute_script(
                "window.scrollTo(0, document.body.scrollHeight * 0.5);"
            )
            time.sleep(2)
            self.driver.execute_script(
                "window.scrollTo(0, document.body.scrollHeight);"
            )
            
            # Espera mais longa para carregar
            time.sleep(3)
            
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                try:
                    load_more = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "next"))
                    )
                    self.driver.execute_script("arguments[0].click();", load_more)
                    time.sleep(3)
                except:
                    break
            
            last_height = new_height
            scroll_attempts += 1
            
            # Verificar se há produtos na página
            products = self.driver.find_elements(By.CSS_SELECTOR, "li.card-product.card-product-showcase, .product-card")
            print(f"Produtos encontrados até agora: {len(products)}")
        
        print("Finished loading all products")

    def cleanup(self):
        """Clean up resources"""
        try:
            if hasattr(self, 'driver') and self.driver:
                self.driver.quit()
        except Exception as e:
            print(f"Error during cleanup: {e}")


    def _extract_links_from_current_page(self):
        """Extract product links from the current page using BeautifulSoup."""
        product_links = set()
        soup = BeautifulSoup(self.driver.page_source, "html.parser")
        
        # Try multiple possible selectors
        product_elements = soup.select("li.card-product.card-product-showcase a, .product-card a")
        
        for element in product_elements:
            if 'href' in element.attrs:
                product_links.add(element['href'])
        
        self.total_products = len(product_links)
        print(f"Extracted {len(product_links)} unique product links")
        
        return product_links

    def _process_product_in_new_instance(self, product_link):
        """Process a single product page and its variations."""
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
        options.add_argument("--no-sandbox")  # Bypass OS security model
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")  # Set user-agent
        driver = webdriver.Chrome(options=options)
        stock_variation = "N/A"  # Default value
        variation_name = "N/A" 
        if self.stop_event.is_set():
            print(f"Stopping processing of product: {product_link}")
            driver.quit()
            return

        try:
            driver.get(product_link)
            print(f"\nProcessing product page: {product_link}")

            # Base product data retrieval with increased timeout
            try:
                nome = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "product-title"))
                ).text.strip()

                code = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "product-code"))
                ).text.strip()

                codigoBarras = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,"li:not([class]) span.spec-value"))
                ).text.strip()

                marca = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "blue"))
                ).text.strip()
            
            except TimeoutException as te:
                print(f"Timeout while trying to get product data from {product_link}: {te}")
                driver.save_screenshot('error_screenshot.png')  # Capture screenshot for debugging
                return
            except Exception as e:
                print(f"Error getting base product data: {e}")
                return

            def try_get_price(driver):
                """Helper function to try different price selectors"""
                                
                price_selectors = [
                    ".modal-item-price .variacao-item-preco .badge-and-price b",
                    ".modal-item-price .variacao-item-preco b",
                    ".variacao-item-preco b",
                    ".current-price-left",
                    ".badge-and-price b"
                ]

                
                for selector in price_selectors:
                    try:
                        price_element = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        price = price_element.text.strip()
                        if price and not price.isspace():
                            return price
                    except:
                        continue
                return None
            
            def try_get_original_price(driver):
                """Helper function to try different original price selectors"""

                original_price_selectors = [
                    "div.old-price-right span.old-price",
                    # Adicione outros seletores para preço original se necessário
                ]
                
                for selector in original_price_selectors:
                    try:
                        original_price_element = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        original_price = original_price_element.text.strip()
                        if original_price and not original_price.isspace():
                            return original_price
                    except:
                        continue
                return "Sem promoção"
            
            def try_get_stock(driver):
                """Helper function to try different original price selectors"""

                stock_selector = [
                    "available-product",
                ]
                
                for selector in stock_selector:
                    try:
                        stock_element = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.ID,selector))
                        )
                        stock = stock_element.text.strip()
                        if stock == "Em estoque":
                            return "Em Estoque"
                        else:
                            return "Sem Estoque"
                    except:
                        continue
                return "Sem Estoque"

            def try_get_subscriber_price(driver):
                """Helper function to try different subscriber price selectors"""
                sub_price_selectors = [
                    ".modal-item-price .price-subscriber",
                    ".price-subscriber",
                    ".variacao-item .price-subscriber"
                ]
                
                for selector in sub_price_selectors:
                    try:
                        sub_price_element = WebDriverWait(driver, 2).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        sub_price = sub_price_element.text.strip()
                        if sub_price and not sub_price.isspace():
                            return sub_price
                    except:
                        continue
                return "N/A"

            def try_get_variation_name(driver):
                """Helper function to try different variation name selectors"""
                variation_name_selectors = [
                    ".product-variation .nome-variacao",
                    ".variation .nome-variacao",
                    ".nome-variacao b"
                ]
                
                for selector in variation_name_selectors:
                    try:
                        print(f"Trying selector: {selector}")
                        variation_name_element = WebDriverWait(driver, 3).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        variation_name = variation_name_element.text.strip()
                        if variation_name and not variation_name.isspace():
                            print(f"Found variation name: {variation_name}")
                            return variation_name
                    except Exception as e:
                        print(f"Pass")
                        continue
                return "Única"


            # Check for variations
            has_variations = False
            try:
                size_button = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "size-select-button"))
                )
                has_variations = True
            except TimeoutException:
                has_variations = False

            if has_variations:
                try:
                    def get_fresh_variations():
                        """Helper function to get fresh variation elements"""
                        fresh_button = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "size-select-button"))
                        )
                        driver.execute_script("arguments[0].scrollIntoView(true);", fresh_button)
                        time.sleep(1)
                        driver.execute_script("arguments[0].click();", fresh_button)
                        time.sleep(1)
                        
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.ID, "popupVariacoes"))
                        )
                        return WebDriverWait(driver, 5).until(
                            EC.presence_of_all_elements_located((By.CLASS_NAME, "variacao-item"))
                        )

                    variations = get_fresh_variations()
                    variation_count = len(variations)
                    print(f"Found {variation_count} variations")

                    for index in range(variation_count):
                        if self.stop_event.is_set():
                            break
                            
                        try:
                            if index > 0:
                                variations = get_fresh_variations()
                            
                            # Store variation text for debugging
                            variation_text = variations[index].text
                            print(f"\nProcessing variation {index + 1}: {variation_text}")
                            
                            # Click the variation
                            driver.execute_script("arguments[0].click();", variations[index])
                            time.sleep(1.5)  # Increased wait time
                            
                            # Try to get prices
                            variation_price = try_get_price(driver)
                            original_variation = try_get_original_price(driver)
                            stock_variation = try_get_stock(driver)
                            variation_name = try_get_variation_name(driver)
                            
                            if not variation_price:
                                print(f"Could not find price for variation {index + 1}")
                                # Try one more time after a longer wait
                                time.sleep(1)
                                variation_price = try_get_price(driver)
                                if not variation_price:
                                    continue
                            
                            if not original_variation:
                                print(f"Could not find price for variation {index + 1}")
                                # Try one more time after a longer wait
                                time.sleep(1)
                                original_variation = try_get_original_price(driver)
                                if not original_variation:
                                    continue

                            if not stock_variation:
                                print(f"Could not find price for variation {index + 1}")
                                # Try one more time after a longer wait
                                time.sleep(1)
                                stock_variation = try_get_original_price(driver)
                                if not stock_variation:
                                    continue
                                
                            variation_pricesub = try_get_subscriber_price(driver)
                            
                            # Save variation data
                            variation_data = [nome, code,codigoBarras ,marca, variation_price,original_variation ,variation_pricesub,stock_variation, variation_name ]
                            self._append_to_excel(variation_data)
                            print(f"Saved variation {index + 1}: {variation_data}")
                            
                            # Close modal before next iteration
                            try:
                                close_button = WebDriverWait(driver, 3).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, ".modal .close"))
                                )
                                driver.execute_script("arguments[0].click();", close_button)
                                time.sleep(0.5)
                            except:
                                pass
                                
                        except Exception as e:
                            print(f"Error processing variation {index + 1}: {str(e)}")
                            try:
                                close_button = WebDriverWait(driver, 3).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, ".modal .close"))
                                )
                                driver.execute_script("arguments[0].click();", close_button)
                            except:
                                pass
                            continue
                            
                except Exception as e:
                    print(f"Error processing variations: {e}")
                    
            else:
                # Process single product without variations
                try:
                    variation_name = try_get_variation_name(driver)
                    stock_variation = try_get_stock(driver)
                    price = try_get_price(driver)
                    if not price:
                        print("Could not find price for single product")
                        return
                    
                    originPrice = try_get_original_price(driver)
                    if not originPrice:
                        print("Could not find Original price for single product")
                        return
                        
                    pricesub = try_get_subscriber_price(driver)
                        
                    product_data = [nome, code,codigoBarras, marca, price,originPrice ,pricesub,stock_variation, variation_name]
                    self._append_to_excel(product_data)
                    print(f"Saved single product: {product_data}")
                    
                except Exception as e:
                    print(f"Error saving single product: {e}")

        except Exception as e:
            print(f"Error processing page {product_link}: {e}")
        finally:
            driver.quit()

    def _process_single_product(self, driver):
        """Helper method to process a single product or variation"""
        try:
            # Wait for price element with extended timeout
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "current-price-left"))
            )

            # Get product details with explicit waits
            nome = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "product-title"))
            ).text.strip()
            
            code = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "product-code"))
            ).text.strip()

            codigoBarras = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,"li:not([class]) span.spec-value"))
                ).text.strip()
            
            marca = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "blue"))
            ).text.strip()
            
            price = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "current-price-left"))
            ).text.strip()

            try:
                originalPrice = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.old-price-right span.old-price"))
                ).text.strip()
            except Exception:
                originalPrice = "Sem promoção"

            try:
                pricesub = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "price-subscriber"))
                ).text.strip()
            except TimeoutException:
                pricesub = "N/A"

            product_data = [nome, code,codigoBarras, marca, price,originalPrice ,pricesub]
            print(f"Extracted Product Data: {product_data}")
            self._append_to_excel(product_data)

        except Exception as e:
            print(f"Error extracting product data: {e}")

    def stop_scraping():
        """This function stops the scraping process."""
        print("Stopping the scraping process in cachorro.py...")
        sys.exit()  #

if __name__ == "__main__":
    driver = webdriver.Chrome()
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")


    try:
        booking_filtration = BookingFiltration(driver)
        booking_filtration.filter_products()
    finally:
        driver.quit()
